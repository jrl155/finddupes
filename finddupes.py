#!/usr/bin/python
# -*- coding: utf-8 -*-
# Copyright (C) 2013 John Hampton <pacopablo@pacopablo.com>
# All rights reserved.
#
# This software is licensed as described in the file COPYING, which
# you should have received as part of this distribution.
#
# Author: John Hampton <pacopablo@pacopablo.com>
#
# Script to find duplicate files in a given directory heirarchy.

"""Find Duplicates

Program to find duplicate files in a given file heirarchy.

finddups.py will traverse a file heirarchy and print out, or save to an excel
spreadsheet, a list of files that duplicates of each other.  Multiple
may be specified on the command line separated by spaces.  Use quotes for
direcotries that contain spaces.  If multiple directories are specified and
one directory is a subfodler of another, there will be duplicate entries in
the output.

Usage:
  finddups.py [-hx <file>] <directory>...

Options:
  -h --help               Show this screen
  --version               Show version
  -x <file> --xls=<file>  Write results to the file given in excel format

"""

# Standard library imports
import sys
import hashlib
import pickle
import os
import collections
from datetime import datetime

# Third Party imports
try:
    from openpyxl import Workbook
    XLS_SUPPORT = True
except ImportError:
    # We don't want excel output to be a requirement, but we need to
    # invalidate teh -x flag if it's not available
    XLS_SUPPORT = False
from docopt import docopt

# Local imports



def pprint_size(size):
    """Print suffixes for the given size.

    Pulled from:
    http://www.dzone.com/snippets/filesize-nice-units
    """
    suffixes = [("B",2**10), ("K",2**20), ("M",2**30), ("G",2**40), ("T",2**50)]
    for suf, lim in suffixes:
        if size > lim:
            continue
        else:
            return round(size/float(lim/2**10),2).__str__()+suf


def pprint_timestamp(timestamp, format="%m/%d/%Y %H:%M:%S"):
    """ Print a timestamp in a readable format.

    A format string may be passed in via the format parameter.
    """
    return datetime.fromtimestamp(timestamp).strftime(format)


def lazyprop(fn):
    """ Function decorator to implemnt lazy load of properties.

    Pulled from:
    http://stackoverflow.com/a/3013910
    """
    attr_name = '_lazy_' + fn.__name__

    @property
    def _lazyprop(self):
        if not hasattr(self, attr_name):
            setattr(self, attr_name, fn(self))
        return getattr(self, attr_name)
    return _lazyprop


class FileData(object):
    """ Object containing identifying information on a given file

    A FileObject is passed a file path in its instanciation.  It gathers the
    following information about the given file:

        Path: absolute path of the file given
        Size: file size
        Timestamp: ctime of the file
        Hash: SHA1 hash of file contents.  The hash is a lazily loaded
              attribute since we may not need to hash every file.

    """

    def __init__(self, path):
        self.path = os.path.abspath(path)
        statinfo = os.stat(self.path)
        self.size = statinfo.st_size
        self.mtime = statinfo.st_mtime
        self.ctime = statinfo.st_ctime
        self.timestamp = self.ctime


    @classmethod
    def hashfile(cls, path, chunksize=160*64):
        """ Return a SHA1 hash (hexdigest) of the file specified in `path`

        Taken from:
        http://stackoverflow.com/a/4213255
        """

        h = hashlib.sha1()
        with open(path,'rb') as f:
            for chunk in iter(lambda: f.read(chunksize), b''):
                h.update(chunk)
        return h.hexdigest()


    @lazyprop
    def hash(self):
        return self.hashfile(self.path)


    def __repr__(self):
        return 'FileData(%s)' % self.path


    def __str__(self):
        return '%s\t%s\t%s\t%s\t%s' % (self.path, self.hash,
                                      pprint_size(self.size),
                                      self.ctime, self.mtime)


def duplicates(hashes):
    """Return the value of any keys that have duplicate files"""

    for hash, files in hashes.items():
        if len(files) > 1:
            yield (hash, files)
        continue


def hash_files(filelist):
    """ Hash files of the same size and return a dictionary of hashes """

    hashes = collections.defaultdict(list)
    for size, files in filelist.items():
        if len(files) > 1:
            for f in files:
                hashes[f.hash].append(f)
                continue
        continue
    return hashes


def main(args):

    for dir in args['<directory>']:
        if not os.path.isdir(os.path.abspath(dir)):
            print("All paths given must be directories. Please check that"
                  " the path given exists and is a directory")
            return 1

    excel = True if args['--xls'] and XLS_SUPPORT else False

    if excel:
        wb = Workbook()
        dup_ws = wb.get_active_sheet()
        dup_ws.title = 'Duplicates'
        dup_ws.cell(row=0, column=0).value = 'Path'
        dup_ws.cell(row=0, column=1).value = 'Hash'
        dup_ws.cell(row=0, column=2).value = 'Size'
        dup_ws.cell(row=0, column=3).value = 'ctime'
        dup_ws.cell(row=0, column=4).value = 'mtime'
        zero_ws = wb.create_sheet()
        zero_ws.title = 'Zero-byte files'
        zero_ws.cell(row=0, column=0).value = 'Path'
        zero_ws.cell(row=0, column=1).value = 'Hash'
        zero_ws.cell(row=0, column=3).value = 'ctime'
        zero_ws.cell(row=0, column=4).value = 'mtime'

    filesizes = collections.defaultdict(list)
    for dir in args['<directory>']:
        for root, dirs, files in os.walk(dir):
            for name in files:
                f = FileData(os.path.join(root, name))
                filesizes[f.size].append(f)
            continue

    row = 1
    # Deal with zero-byte files
    num_zero_byte_files = len(filesizes[0])
    if excel:
        for row, filename in enumerate(filesizes[0]):
            zero_ws.cell(row=row+1, column=0).value = f.path
            zero_ws.cell(row=row+1, column=1).value = f.hash
            zero_ws.cell(row=row+1, column=3).value = pprint_timestamp(f.ctime)
            zero_ws.cell(row=row+1, column=4).value = pprint_timestamp(f.mtime)
            continue
    del filesizes[0]

    hashes = hash_files(filesizes)

    row = 1
    for hash, files in duplicates(hashes):
        if not excel:
            print("{}:".format(hash))
        for f in files:
            if excel:
                dup_ws.cell(row=row, column=0).value = f.path
                dup_ws.cell(row=row, column=1).value = f.hash
                dup_ws.cell(row=row, column=2).value = f.size
                dup_ws.cell(row=row, column=3).value = pprint_timestamp(f.ctime)
                dup_ws.cell(row=row, column=4).value = pprint_timestamp(f.mtime)
            else:
                print("{}\t{}\t{}\t{}\t{}".format(pprint_timestamp(f.ctime), pprint_timestamp(f.mtime), pprint_size(f.size), f.hash, f.path))
            row += 1
            continue

        if not excel:
            print("\n")
        continue

    if not excel:
        print("Number of zero byte files: %d" % num_zero_byte_files)
    else:
        wb.save(args['--xls'])

    return 0


if __name__ == '__main__':

    args = docopt(__doc__, version='%s 1.0' % os.path.basename(sys.argv[0]))
    sys.exit(main(args))



